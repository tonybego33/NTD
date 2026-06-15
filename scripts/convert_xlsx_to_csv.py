"""
Convertit la feuille EXPORT du fichier Indicateurs_GT_BDDe.xlsx en CSV léger.

Utilisation :
    python scripts/convert_xlsx_to_csv.py <chemin_vers_fichier_xlsx>

Le CSV sera écrit dans backend/data/indicateurs_export.csv.

Le CSV est utilisable par le service backend.services.indicateurs_locaux.
"""
from __future__ import annotations

import csv
import sys
from pathlib import Path

import openpyxl


def convert(xlsx_path: Path, csv_path: Path) -> None:
    print(f"Lecture : {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if "EXPORT" not in wb.sheetnames:
        raise SystemExit(f"Feuille 'EXPORT' introuvable. Feuilles : {wb.sheetnames}")
    ws = wb["EXPORT"]

    csv_path.parent.mkdir(parents=True, exist_ok=True)

    n = 0
    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            # Nettoyer les None → ""
            cleaned = [("" if v is None else v) for v in row]
            writer.writerow(cleaned)
            n += 1
            if n % 5000 == 0:
                print(f"  ... {n} lignes écrites")
    wb.close()

    size_kb = csv_path.stat().st_size / 1024
    print(f"\n✓ Ecrit : {csv_path} ({n} lignes, {size_kb:.0f} Ko)")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage : python scripts/convert_xlsx_to_csv.py <fichier_xlsx>")
        sys.exit(1)
    xlsx = Path(sys.argv[1]).resolve()
    if not xlsx.exists():
        raise SystemExit(f"Fichier introuvable : {xlsx}")
    # Le CSV va dans backend/data/ relatif à la racine du projet
    project_root = Path(__file__).resolve().parent.parent
    csv_path = project_root / "backend" / "data" / "indicateurs_export.csv"
    convert(xlsx, csv_path)
