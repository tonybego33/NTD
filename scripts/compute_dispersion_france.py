"""
Lit un ZIP INSEE Filosofi historique (contient un XLSX multi-onglets),
extrait CODGEO + médiane revenu disponible par UC, génère un mini-CSV
consommable par precompute_cache.py.

INSEE distribue les Filosofi 2017 et 2019 sous forme de XLSX zippé :
  - indic-struct-distrib-revenu-2017-COMMUNES.zip
  - indic-struct-distrib-revenu-2019-COMMUNES.zip

Le XLSX contient plusieurs onglets ; on cherche celui qui contient CODGEO + MEDxx
(le niveau de vie médian par unité de consommation).

Usage :
  python -m scripts.convert_filosofi_xlsx <chemin_zip_ou_xlsx> <année>
"""
from __future__ import annotations

import csv
import io
import sys
import zipfile
from pathlib import Path


def find_xlsx_in_zip(zip_path: Path) -> str | None:
    with zipfile.ZipFile(zip_path, "r") as z:
        for name in z.namelist():
            if name.lower().endswith((".xlsx", ".xls")):
                return name
    return None


def load_xlsx_buffer(path: Path) -> io.BytesIO:
    """Retourne un buffer BytesIO contenant le XLSX (depuis un ZIP ou un fichier direct)."""
    if path.suffix.lower() == ".zip":
        xlsx_name = find_xlsx_in_zip(path)
        if not xlsx_name:
            raise FileNotFoundError(f"Aucun XLSX dans {path.name}")
        print(f"  XLSX trouvé dans le ZIP : {xlsx_name}")
        with zipfile.ZipFile(path, "r") as z:
            with z.open(xlsx_name) as f:
                return io.BytesIO(f.read())
    elif path.suffix.lower() in (".xlsx", ".xls"):
        with open(path, "rb") as f:
            return io.BytesIO(f.read())
    raise ValueError(f"Format non supporté : {path.suffix}")


def find_median_sheet(wb, year: int):
    """Cherche l'onglet contenant CODGEO + MEDxx dans son en-tête."""
    yy = str(year)[2:]
    target_med = f"MED{yy}"

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Lit les 20 premières lignes pour trouver l'en-tête
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
            cells = [str(c).strip().upper() if c is not None else "" for c in row]
            if "CODGEO" in cells:
                # Vérifie qu'on a aussi une colonne médiane (MEDxx, MED, MEDIANE...)
                has_med = any(
                    c == target_med or c == "MED" or "MEDIANE" in c or
                    (c.startswith("MED") and len(c) == 5)
                    for c in cells
                )
                if has_med:
                    return ws, row_idx, cells
    return None, None, None


def extract_medianes(buf: io.BytesIO, year: int) -> list[tuple[str, float]]:
    """Lit le XLSX, trouve le bon onglet, extrait CODGEO + MED."""
    try:
        import openpyxl
    except ImportError:
        sys.exit(
            "[err] openpyxl manquant. Installe-le :\n"
            "      pip install openpyxl --break-system-packages"
        )

    wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
    print(f"  Onglets disponibles : {wb.sheetnames}")

    ws, header_row, headers = find_median_sheet(wb, year)
    if ws is None:
        raise ValueError(
            f"Aucun onglet ne contient CODGEO + MED{str(year)[2:]}.\n"
            f"Onglets vus : {wb.sheetnames}"
        )

    print(f"  Onglet retenu : '{ws.title}' (en-tête ligne {header_row})")

    # Trouve les indices des colonnes
    cg_idx = headers.index("CODGEO")
    yy = str(year)[2:]
    med_idx = None
    candidates = [f"MED{yy}", "MED", "MEDIANE"]
    for cand in candidates:
        if cand in headers:
            med_idx = headers.index(cand)
            break
    if med_idx is None:
        # Fallback : 1ère colonne qui commence par MED
        for i, h in enumerate(headers):
            if h.startswith("MED"):
                med_idx = i
                break
    if med_idx is None:
        raise ValueError(f"Colonne médiane introuvable. Headers : {headers[:15]}")

    print(f"  Colonne médiane : '{headers[med_idx]}'")

    # Lit les données
    data = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if cg_idx >= len(row) or med_idx >= len(row):
            continue
        cg = row[cg_idx]
        med = row[med_idx]
        if cg is None or med is None:
            continue
        cg_str = str(cg).strip()
        # Filtre : communes (5 chars, peut commencer par 0 ou 2A/2B pour Corse)
        # ou EPCI (9 chars numériques)
        if len(cg_str) == 5:
            pass  # commune
        elif len(cg_str) == 9 and cg_str.isdigit():
            pass  # EPCI
        else:
            continue
        # Conversion de la valeur
        try:
            med_float = float(str(med).replace(",", ".").replace(" ", "").replace("\xa0", ""))
        except (ValueError, TypeError):
            continue
        if med_float <= 0:
            continue
        data.append((cg_str, med_float))

    return data


def main():
    if len(sys.argv) < 3:
        sys.exit("Usage : python -m scripts.convert_filosofi_xlsx <chemin_zip_ou_xlsx> <année>")

    src = Path(sys.argv[1])
    year = int(sys.argv[2])

    if not src.exists():
        sys.exit(f"[err] Fichier introuvable : {src}")

    print(f"Lecture {src.name} (Filosofi {year})...")
    buf = load_xlsx_buffer(src)
    data = extract_medianes(buf, year)

    n_com = sum(1 for cg, _ in data if len(cg) == 5)
    n_epci = sum(1 for cg, _ in data if len(cg) == 9)
    print(f"  → {n_com:,} communes + {n_epci:,} EPCI extraites".replace(",", " "))

    out = Path(__file__).resolve().parent.parent / "backend" / "data" / f"filosofi_revenu_{year}.csv"
    out.parent.mkdir(parents=True, exist_ok=True)

    with open(out, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["CODGEO", f"revenu_median_{year}"])
        for cg, med in sorted(data):
            w.writerow([cg, med])

    print(f"\n✓ Sortie : {out}")
    print(f"  Taille : {out.stat().st_size / 1024:.1f} Ko")


if __name__ == "__main__":
    main()
